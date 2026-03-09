"""Tests for events list, export, and deduplication."""

from datetime import date, timedelta
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.event import Event
from app.models.legislator import Legislator
from app.scraper.event_data import EventData
from app.scraper.orchestrator import ScraperOrchestrator


@pytest.fixture
async def future_events(db_session: AsyncSession, sample_legislator: Legislator):
    """Create a mix of future and past events (committed for client visibility)."""
    today = date.today()
    future = (today + timedelta(days=7)).isoformat()
    past = (today - timedelta(days=7)).isoformat()

    events = [
        Event(
            legislator_id=sample_legislator.id,
            title="Future Town Hall",
            date=future,
            time="18:00",
            address="123 Main St",
            event_type="town_hall",
        ),
        Event(
            legislator_id=sample_legislator.id,
            title="Past Town Hall",
            date=past,
            time="18:00",
            address="456 Oak Ave",
            event_type="town_hall",
        ),
        Event(
            legislator_id=sample_legislator.id,
            title="Undated Event",
            date=None,
            time=None,
            address="789 Pine Rd",
            event_type="community_meeting",
        ),
    ]
    for ev in events:
        db_session.add(ev)
    await db_session.commit()
    return events


class TestListEvents:
    async def test_returns_future_events_only(self, client, auth_headers, future_events):
        resp = await client.get("/api/events", headers=auth_headers)
        assert resp.status_code == 200
        body = resp.json()
        assert "events" in body
        assert "total" in body
        assert "page" in body
        titles = [e["title"] for e in body["events"]]
        assert "Future Town Hall" in titles
        assert "Undated Event" in titles  # null dates included
        assert "Past Town Hall" not in titles

    async def test_filter_by_chamber(self, client, auth_headers, future_events):
        resp = await client.get("/api/events?chamber=assembly", headers=auth_headers)
        assert resp.status_code == 200
        body = resp.json()
        for ev in body["events"]:
            assert ev["legislator_chamber"] == "assembly"

    async def test_filter_by_search(self, client, auth_headers, future_events):
        resp = await client.get("/api/events?search=Undated", headers=auth_headers)
        assert resp.status_code == 200
        body = resp.json()
        assert body["total"] == 1
        assert body["events"][0]["title"] == "Undated Event"

    async def test_filter_by_date_range(self, client, auth_headers, future_events):
        today = date.today()
        start = (today + timedelta(days=5)).isoformat()
        end = (today + timedelta(days=10)).isoformat()
        resp = await client.get(
            f"/api/events?start_date={start}&end_date={end}",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        body = resp.json()
        assert body["total"] == 1
        assert body["events"][0]["title"] == "Future Town Hall"

    async def test_pagination(self, client, auth_headers, future_events):
        resp = await client.get("/api/events?per_page=1&page=1", headers=auth_headers)
        assert resp.status_code == 200
        body = resp.json()
        assert len(body["events"]) == 1
        assert body["per_page"] == 1
        assert body["total"] >= 1
        assert body["total_pages"] >= 1

    async def test_requires_auth(self, client):
        resp = await client.get("/api/events")
        assert resp.status_code == 403


class TestExportEvents:
    async def test_returns_xlsx(self, client, auth_headers, future_events):
        resp = await client.get("/api/events/export", headers=auth_headers)
        assert resp.status_code == 200
        assert (
            resp.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    async def test_xlsx_headers(self, client, auth_headers, future_events):
        resp = await client.get("/api/events/export", headers=auth_headers)
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert headers == [
            "NAME",
            "DATE",
            "TIME",
            "ADDRESS",
            "TITLE OF EVENT",
            "EVENT LINK",
            "ADDITIONAL DETAILS",
        ]

    async def test_xlsx_legislator_name_format(self, client, auth_headers, future_events):
        resp = await client.get("/api/events/export", headers=auth_headers)
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        # Row 2 is first data row
        if ws.max_row >= 2:
            name = ws.cell(row=2, column=1).value
            assert "Assemblymember" in name
            assert "Jane Smith" in name
            assert "(D-42)" in name

    async def test_xlsx_date_formatted(self, client, auth_headers, future_events):
        resp = await client.get("/api/events/export", headers=auth_headers)
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        # Find a row with a non-empty date (skip header row 1 and any null-date rows)
        date_val = None
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=2).value
            if val:
                date_val = val
                break
        assert date_val is not None
        assert "-" not in str(date_val)  # not ISO format
        assert ", " in str(date_val)  # has comma like "Month DD, YYYY"

    async def test_xlsx_time_formatted(self, client, auth_headers, future_events):
        resp = await client.get("/api/events/export", headers=auth_headers)
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        # Find a row with a non-empty time
        time_val = None
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=3).value
            if val:
                time_val = val
                break
        assert time_val is not None
        assert "PM" in str(time_val) or "AM" in str(time_val)


class TestDeduplication:
    async def test_upsert_same_title_date_updates(
        self, db_session: AsyncSession, sample_legislator: Legislator
    ):
        orch = ScraperOrchestrator()
        events_v1 = [
            EventData(
                title="Town Hall",
                date="2026-06-01",
                time="18:00",
                address="Old Address",
            )
        ]
        await orch._save_events(events_v1, sample_legislator.id, db_session)
        await db_session.commit()

        # Save again with updated address
        events_v2 = [
            EventData(
                title="Town Hall",
                date="2026-06-01",
                time="19:00",
                address="New Address",
            )
        ]
        await orch._save_events(events_v2, sample_legislator.id, db_session)
        await db_session.commit()

        result = await db_session.execute(
            select(Event).where(
                Event.legislator_id == sample_legislator.id,
                Event.title == "Town Hall",
                Event.date == "2026-06-01",
            )
        )
        all_events = result.scalars().all()
        assert len(all_events) == 1
        assert all_events[0].address == "New Address"
        assert all_events[0].time == "19:00"

    async def test_different_dates_creates_separate(
        self, db_session: AsyncSession, sample_legislator: Legislator
    ):
        orch = ScraperOrchestrator()
        events = [
            EventData(title="Town Hall", date="2026-06-01"),
            EventData(title="Town Hall", date="2026-06-15"),
        ]
        await orch._save_events(events, sample_legislator.id, db_session)
        await db_session.commit()

        result = await db_session.execute(
            select(Event).where(
                Event.legislator_id == sample_legislator.id,
                Event.title == "Town Hall",
            )
        )
        all_events = result.scalars().all()
        assert len(all_events) == 2

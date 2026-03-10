import io
import logging
import os as _os
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo

from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from fastapi import Depends, FastAPI, HTTPException, Query, Request
from fastapi.exceptions import RequestValidationError
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from pydantic import BaseModel, ConfigDict, EmailStr, Field, field_validator
from sqlalchemy import Integer, cast, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.auth import (
    create_access_token,
    create_refresh_token,
    get_current_user,
    hash_password,
    revoke_all_user_tokens,
    rotate_refresh_token,
    verify_password,
    verify_password_timing_safe,
)
from app.config import ALLOWED_ORIGINS, SCRAPE_CRON, SCRAPE_ENABLED
from app.database import get_session
from app.invite import generate_invite_code, validate_and_consume_invite
from app.models.event import Event
from app.models.legislator import Legislator
from app.models.scrape_log import ScrapeLog
from app.models.user import User
from app.password_check import is_breached_password
from app.rate_limit import limiter
from app.request_context import generate_request_id, mask_email, request_id_var
from app.scrape_runner import get_job, run_full_scrape, run_single_scrape

logger = logging.getLogger(__name__)
_security_logger = logging.getLogger("security")

PACIFIC = ZoneInfo("America/Los_Angeles")


def _log_security_event(
    event_type: str,
    outcome: str,
    ip: str,
    email: str | None = None,
    reason: str | None = None,
):
    """Emit a structured security log entry to the 'security' logger."""
    extra = {
        "event_type": event_type,
        "outcome": outcome,
        "ip": ip,
    }
    if email:
        extra["email_masked"] = mask_email(email)
    if reason:
        extra["reason"] = reason

    level = logging.INFO if outcome == "success" else logging.WARNING
    _security_logger.log(level, "%s: %s", event_type, outcome, extra=extra)


def _to_pacific_iso(dt_val: datetime | None) -> str | None:
    """Convert a naive-UTC datetime to a Pacific-timezone ISO string."""
    if dt_val is None:
        return None
    aware = dt_val.replace(tzinfo=timezone.utc)
    return aware.astimezone(PACIFIC).isoformat()


def _today_pacific() -> str:
    """Return today's date in Pacific timezone as YYYY-MM-DD."""
    return datetime.now(PACIFIC).strftime("%Y-%m-%d")


scheduler = AsyncIOScheduler()


async def scheduled_scrape():
    logger.info("Scheduled scrape triggered")
    await run_full_scrape()


async def _seed_bootstrap_invite():
    """Seed a bootstrap invite code from INVITE_CODE env var (idempotent).

    This allows the first deployment to work with a known invite code.
    The code gets 100 uses and 365 days expiry. Once the app is running,
    use the /api/auth/invite-codes endpoint to generate proper codes.
    """
    import hashlib
    import os

    from app.database import async_session
    from app.models.invite_code import InviteCode

    bootstrap_code = os.getenv("INVITE_CODE", "")
    if not bootstrap_code:
        return

    code_hash = hashlib.sha256(bootstrap_code.encode()).hexdigest()
    async with async_session() as session:
        result = await session.execute(select(InviteCode).where(InviteCode.code_hash == code_hash))
        if result.scalar_one_or_none():
            return  # Already seeded

        invite = InviteCode(
            code_hash=code_hash,
            max_uses=100,
            times_used=0,
            expires_at=datetime.utcnow() + timedelta(days=365),
            created_by=None,
        )
        session.add(invite)
        await session.commit()
        logger.info("Seeded bootstrap invite code from INVITE_CODE env var")


@asynccontextmanager
async def lifespan(app: FastAPI):
    from app.logging_config import setup_logging

    setup_logging()

    # Auto-seed legislators on startup (idempotent)
    from scripts.seed_legislators import seed_legislators

    result = await seed_legislators()
    logger.info("Legislator seed: %d created, %d updated", result["created"], result["updated"])

    # Seed bootstrap invite code from env var (idempotent)
    await _seed_bootstrap_invite()

    if SCRAPE_ENABLED and SCRAPE_CRON:
        parts = SCRAPE_CRON.split()
        if len(parts) == 5:
            trigger = CronTrigger(
                minute=parts[0],
                hour=parts[1],
                day=parts[2],
                month=parts[3],
                day_of_week=parts[4],
                timezone="US/Pacific",
            )
            scheduler.add_job(scheduled_scrape, trigger, id="daily_scrape", replace_existing=True)
            scheduler.start()
            logger.info("Scheduled scrape: %s Pacific", SCRAPE_CRON)
    yield
    if scheduler.running:
        scheduler.shutdown(wait=False)


_is_production = _os.getenv("RAILWAY_ENVIRONMENT") or _os.getenv("NODE_ENV") == "production"

app = FastAPI(
    title="CA Town Hall Monitor",
    lifespan=lifespan,
    # Disable API docs in production
    docs_url=None if _is_production else "/docs",
    redoc_url=None if _is_production else "/redoc",
    openapi_url=None if _is_production else "/openapi.json",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["Strict-Transport-Security"] = "max-age=63072000; includeSubDomains; preload"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    if "server" in response.headers:
        del response.headers["server"]
    return response


@app.middleware("http")
async def request_id_middleware(request: Request, call_next):
    """Assign a unique request ID to every request for log traceability."""
    rid = generate_request_id()
    token = request_id_var.set(rid)
    response = await call_next(request)
    response.headers["X-Request-ID"] = rid
    request_id_var.reset(token)
    return response


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    """Return generic 422 without exposing field-level schema details."""
    return JSONResponse(status_code=422, content={"detail": "Invalid request data"})


@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    """Catch-all: log the real error, return generic 500."""
    logger.exception("Unhandled exception on %s %s", request.method, request.url.path)
    return JSONResponse(status_code=500, content={"detail": "Internal server error"})


# ---------------------------------------------------------------------------
# Health (public)
# ---------------------------------------------------------------------------


@app.get("/api/health")
async def health():
    return {"status": "ok"}


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------


class StrictBase(BaseModel):
    """Base model that rejects unexpected fields."""

    model_config = ConfigDict(extra="forbid")


class LoginRequest(StrictBase):
    email: EmailStr = Field(max_length=254)
    password: str = Field(min_length=1, max_length=128)

    @field_validator("email", mode="before")
    @classmethod
    def normalize_email(cls, v: str) -> str:
        return v.strip().lower()


class RegisterRequest(StrictBase):
    email: EmailStr = Field(max_length=254)
    password: str = Field(min_length=10, max_length=128)
    name: str = Field(min_length=1, max_length=100)
    invite_code: str = Field(min_length=1, max_length=64)

    @field_validator("email", mode="before")
    @classmethod
    def normalize_email(cls, v: str) -> str:
        return v.strip().lower()

    @field_validator("name", mode="before")
    @classmethod
    def strip_name(cls, v: str) -> str:
        return v.strip()

    @field_validator("invite_code", mode="before")
    @classmethod
    def strip_invite_code(cls, v: str) -> str:
        return v.strip()

    @field_validator("password")
    @classmethod
    def check_breached(cls, v: str) -> str:
        if is_breached_password(v):
            raise ValueError("This password is too common and has appeared in data breaches")
        return v


@app.post("/api/auth/login")
async def login(body: LoginRequest, request: Request, session: AsyncSession = Depends(get_session)):
    ip = request.headers.get(
        "x-forwarded-for", request.client.host if request.client else "unknown"
    )
    ip = ip.split(",")[0].strip()

    # Per-IP rate limit
    retry = limiter.check_login_ip(ip)
    if retry:
        _log_security_event("login_rate_limit", "blocked", ip, body.email, reason="ip_limit")
        raise HTTPException(
            status_code=429, detail="Too many login attempts", headers={"Retry-After": str(retry)}
        )

    # Per-account lock check
    retry = limiter.check_account_lock(body.email)
    if retry:
        _log_security_event("login_rate_limit", "blocked", ip, body.email, reason="account_locked")
        raise HTTPException(
            status_code=429, detail="Too many login attempts", headers={"Retry-After": str(retry)}
        )

    limiter.record_login_ip(ip)

    result = await session.execute(select(User).where(User.email == body.email))
    user = result.scalar_one_or_none()
    # Always run bcrypt even if user is None — prevents timing side-channel
    if not verify_password_timing_safe(body.password, user):
        limiter.record_failed_login(body.email)
        _log_security_event("login", "failure", ip, body.email)
        raise HTTPException(status_code=401, detail="Invalid email or password")

    limiter.clear_failed_logins(body.email)
    _log_security_event("login", "success", ip, body.email)
    access_token = create_access_token(user.id, user.email, user.token_version)
    refresh_token = await create_refresh_token(user.id, session)
    return {
        "token": access_token,
        "refresh_token": refresh_token,
        "user": {"id": user.id, "email": user.email, "name": user.name},
    }


@app.post("/api/auth/register")
async def register(
    body: RegisterRequest, request: Request, session: AsyncSession = Depends(get_session)
):
    ip = request.headers.get(
        "x-forwarded-for", request.client.host if request.client else "unknown"
    )
    ip = ip.split(",")[0].strip()

    # Per-IP invite/register rate limit
    retry = limiter.check_register_ip(ip)
    if retry:
        _log_security_event("register_rate_limit", "blocked", ip, body.email)
        raise HTTPException(
            status_code=429,
            detail="Too many registration attempts",
            headers={"Retry-After": str(retry)},
        )

    limiter.record_register_ip(ip)

    if not await validate_and_consume_invite(body.invite_code, ip, session):
        _log_security_event("register", "failure", ip, body.email, reason="invalid_invite")
        raise HTTPException(status_code=403, detail="Registration failed")

    existing = await session.execute(select(User).where(User.email == body.email))
    if existing.scalar_one_or_none():
        # Same generic error as invalid invite — no way to distinguish
        _log_security_event("register", "failure", ip, body.email, reason="duplicate_email")
        raise HTTPException(status_code=403, detail="Registration failed")

    user = User(
        email=body.email,
        hashed_password=hash_password(body.password),
        name=body.name,
    )
    session.add(user)
    await session.commit()
    await session.refresh(user)

    _log_security_event("register", "success", ip, body.email)
    access_token = create_access_token(user.id, user.email, user.token_version)
    refresh_token = await create_refresh_token(user.id, session)
    return {
        "token": access_token,
        "refresh_token": refresh_token,
        "user": {"id": user.id, "email": user.email, "name": user.name},
    }


class RefreshRequest(StrictBase):
    refresh_token: str = Field(min_length=1, max_length=128)


@app.post("/api/auth/refresh")
async def refresh(body: RefreshRequest, session: AsyncSession = Depends(get_session)):
    result = await rotate_refresh_token(body.refresh_token, session)
    if not result:
        raise HTTPException(status_code=401, detail="Invalid or expired refresh token")
    return {
        "token": result["access_token"],
        "refresh_token": result["refresh_token"],
        "user": result["user"],
    }


class ChangePasswordRequest(StrictBase):
    current_password: str = Field(min_length=1, max_length=128)
    new_password: str = Field(min_length=10, max_length=128)

    @field_validator("new_password")
    @classmethod
    def check_breached(cls, v: str) -> str:
        if is_breached_password(v):
            raise ValueError("This password is too common and has appeared in data breaches")
        return v


@app.post("/api/auth/change-password")
async def change_password(
    body: ChangePasswordRequest,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    if not verify_password(body.current_password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Current password is incorrect")

    user.hashed_password = hash_password(body.new_password)
    user.token_version += 1  # Invalidate all existing tokens
    await revoke_all_user_tokens(user.id, session)
    await session.commit()
    return {"message": "Password changed. Please log in again."}


@app.post("/api/auth/logout")
async def logout(
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    await revoke_all_user_tokens(user.id, session)
    return {"message": "Logged out"}


@app.get("/api/auth/me")
async def me(user: User = Depends(get_current_user)):
    return {"id": user.id, "email": user.email, "name": user.name}


class CreateInviteRequest(StrictBase):
    max_uses: int = Field(default=1, ge=1, le=100)
    expiry_days: int = Field(default=7, ge=1, le=90)


@app.post("/api/auth/invite-codes")
async def create_invite_code(
    body: CreateInviteRequest,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    code = await generate_invite_code(
        session,
        created_by=user.id,
        max_uses=body.max_uses,
        expiry_days=body.expiry_days,
    )
    return {"invite_code": code, "max_uses": body.max_uses, "expiry_days": body.expiry_days}


# ---------------------------------------------------------------------------
# Legislators (protected)
# ---------------------------------------------------------------------------


@app.post("/api/legislators/seed")
async def seed_legislators_endpoint(_user: User = Depends(get_current_user)):
    from scripts.seed_legislators import seed_legislators

    return await seed_legislators()


@app.get("/api/legislators")
async def list_legislators(
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    result = await session.execute(
        select(Legislator).order_by(Legislator.chamber, cast(Legislator.district, Integer))
    )
    legislators = result.scalars().all()
    return [_legislator_dict(leg) for leg in legislators]


def _legislator_dict(leg: Legislator) -> dict:
    facebook_flag = leg.facebook_url is not None and leg.scrape_status in (
        None,
        "no_events",
        "failed",
    )
    return {
        "id": leg.id,
        "name": leg.name,
        "chamber": leg.chamber,
        "district": leg.district,
        "party": leg.party,
        "official_website": leg.official_website,
        "campaign_website": leg.campaign_website,
        "facebook_url": leg.facebook_url,
        "last_scraped_at": _to_pacific_iso(leg.last_scraped_at),
        "scrape_status": leg.scrape_status,
        "facebook_flag": facebook_flag,
        "facebook_note": ("Events may only be posted on Facebook" if facebook_flag else None),
    }


# ---------------------------------------------------------------------------
# Scrape: run / status / logs (protected)
# ---------------------------------------------------------------------------


@app.post("/api/scrape/run")
async def trigger_scrape(_user: User = Depends(get_current_user)):
    job_id = await run_full_scrape()
    return {"job_id": job_id, "status": "running"}


@app.post("/api/scrape/run/{legislator_id}")
async def trigger_single_scrape(
    legislator_id: int,
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    leg = await session.get(Legislator, legislator_id)
    if not leg:
        raise HTTPException(status_code=404, detail="Legislator not found")
    job_id = await run_single_scrape(legislator_id)
    return {"job_id": job_id, "status": "running", "legislator": leg.name}


@app.get("/api/scrape/status/{job_id}")
async def scrape_status(job_id: str, _user: User = Depends(get_current_user)):
    job = get_job(job_id)
    if not job:
        return {"error": "Job not found or expired"}
    return {
        "id": job["id"],
        "status": job["status"],
        "started_at": _to_pacific_iso(job["started_at"]),
        "completed_at": _to_pacific_iso(job["completed_at"]),
        "total": job["total"],
        "completed_count": job["completed_count"],
        "success": job["success"],
        "no_events": job["no_events"],
        "failed": job["failed"],
        "ai_used": job["ai_used"],
        "ai_total_cost_usd": round(job["ai_total_cost"], 4),
        "past_events_removed": job["past_events_removed"],
        # Indicate failure without exposing internal exception details
        "has_error": bool(job.get("error")),
    }


@app.get("/api/scrape/logs")
async def scrape_logs(
    limit: int = Query(50, ge=1, le=500),
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    logs_q = (
        select(ScrapeLog)
        .options(joinedload(ScrapeLog.legislator))
        .order_by(ScrapeLog.completed_at.desc())
        .limit(limit)
    )
    logs_result = await session.execute(logs_q)
    logs = logs_result.scalars().unique().all()

    stats_q = select(
        func.count().label("total"),
        func.count().filter(ScrapeLog.status == "success").label("success"),
        func.count().filter(ScrapeLog.status == "no_events").label("no_events"),
        func.count().filter(ScrapeLog.status == "failed").label("failed"),
        func.count().filter(ScrapeLog.method_used == "ai").label("ai_used"),
    )
    stats_result = await session.execute(stats_q)
    stats = stats_result.one()

    return {
        "summary": {
            "total_logs": stats.total,
            "success": stats.success,
            "no_events": stats.no_events,
            "failed": stats.failed,
            "ai_used": stats.ai_used,
        },
        "logs": [
            {
                "id": log.id,
                "legislator_name": log.legislator.name if log.legislator else None,
                "legislator_id": log.legislator_id,
                "started_at": _to_pacific_iso(log.started_at),
                "completed_at": _to_pacific_iso(log.completed_at),
                "status": log.status,
                "method_used": log.method_used,
                "has_error": bool(log.error_message),
            }
            for log in logs
        ],
    }


# ---------------------------------------------------------------------------
# Scrape: failures (protected)
# ---------------------------------------------------------------------------


@app.get("/api/scrape/failures")
async def scrape_failures(
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    result = await session.execute(
        select(Legislator)
        .where(Legislator.consecutive_failures >= 3)
        .order_by(Legislator.consecutive_failures.desc())
    )
    legislators = result.scalars().all()

    failures = []
    for leg in legislators:
        # Get the most recent scrape log for error context
        log_result = await session.execute(
            select(ScrapeLog)
            .where(ScrapeLog.legislator_id == leg.id)
            .order_by(ScrapeLog.completed_at.desc())
            .limit(1)
        )
        last_log = log_result.scalar_one_or_none()

        failures.append(
            {
                "id": leg.id,
                "name": leg.name,
                "chamber": leg.chamber,
                "district": leg.district,
                "official_website": leg.official_website,
                "consecutive_failures": leg.consecutive_failures,
                "circuit_open_until": _to_pacific_iso(leg.circuit_open_until),
                "has_error": bool(last_log.error_message) if last_log else False,
                "last_attempt": _to_pacific_iso(last_log.completed_at) if last_log else None,
            }
        )

    return failures


# ---------------------------------------------------------------------------
# Scrape: summary (protected)
# ---------------------------------------------------------------------------


@app.get("/api/scrape/summary")
async def scrape_summary(
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    # Get the most recent completed scrape log to determine last scrape time
    last_log_q = select(ScrapeLog).order_by(ScrapeLog.completed_at.desc()).limit(1)
    last_log_result = await session.execute(last_log_q)
    last_log = last_log_result.scalar_one_or_none()

    last_scrape_time = None
    duration_seconds = None
    if last_log and last_log.completed_at:
        last_scrape_time = _to_pacific_iso(last_log.completed_at)
        if last_log.started_at:
            duration_seconds = int((last_log.completed_at - last_log.started_at).total_seconds())

    # Count legislators
    leg_count_q = select(func.count()).select_from(Legislator)
    total_legislators = (await session.execute(leg_count_q)).scalar() or 0

    # Stats from most recent scrape logs (one per legislator)
    # Use a subquery to get the latest log per legislator
    latest_per_leg = (
        select(
            ScrapeLog.legislator_id,
            func.max(ScrapeLog.completed_at).label("max_completed"),
        )
        .group_by(ScrapeLog.legislator_id)
        .subquery()
    )

    latest_logs_q = select(ScrapeLog).join(
        latest_per_leg,
        (ScrapeLog.legislator_id == latest_per_leg.c.legislator_id)
        & (ScrapeLog.completed_at == latest_per_leg.c.max_completed),
    )
    latest_result = await session.execute(latest_logs_q)
    latest_logs = latest_result.scalars().all()

    success_count = sum(1 for sl in latest_logs if sl.status == "success")
    no_events_count = sum(1 for sl in latest_logs if sl.status == "no_events")
    failed_count = sum(1 for sl in latest_logs if sl.status == "failed")
    skipped_count = sum(1 for sl in latest_logs if sl.status == "skipped")
    ai_count = sum(1 for sl in latest_logs if sl.method_used == "ai")

    # Get legislators for chamber info
    leg_result = await session.execute(select(Legislator))
    all_legs = {lg.id: lg for lg in leg_result.scalars().all()}

    def _chamber_count(status: str, chamber: str) -> int:
        return sum(
            1
            for sl in latest_logs
            if sl.status == status
            and sl.legislator_id in all_legs
            and all_legs[sl.legislator_id].chamber == chamber
        )

    assembly_success = _chamber_count("success", "assembly")
    assembly_failed = _chamber_count("failed", "assembly")
    senate_success = _chamber_count("success", "senate")
    senate_failed = _chamber_count("failed", "senate")

    # Failures (3+ consecutive)
    fail_result = await session.execute(
        select(Legislator)
        .where(Legislator.consecutive_failures >= 3)
        .order_by(Legislator.consecutive_failures.desc())
    )
    problem_legislators = fail_result.scalars().all()

    problem_list = []
    for leg in problem_legislators:
        log_r = await session.execute(
            select(ScrapeLog)
            .where(ScrapeLog.legislator_id == leg.id)
            .order_by(ScrapeLog.completed_at.desc())
            .limit(1)
        )
        last = log_r.scalar_one_or_none()
        problem_list.append(
            {
                "id": leg.id,
                "name": leg.name,
                "chamber": leg.chamber,
                "district": leg.district,
                "consecutive_failures": leg.consecutive_failures,
                "has_error": bool(last.error_message) if last else False,
                "last_attempt": _to_pacific_iso(last.completed_at) if last else None,
            }
        )

    return {
        "last_scrape_time": last_scrape_time,
        "duration_seconds": duration_seconds,
        "total_legislators": total_legislators,
        "success": success_count,
        "no_events": no_events_count,
        "failed": failed_count,
        "skipped": skipped_count,
        "ai_used": ai_count,
        "chamber_breakdown": {
            "assembly": {"success": assembly_success, "failed": assembly_failed},
            "senate": {"success": senate_success, "failed": senate_failed},
        },
        "problem_legislators": problem_list,
    }


# ---------------------------------------------------------------------------
# Events: list / export (protected)
# ---------------------------------------------------------------------------


def _build_events_query(
    chamber: str | None,
    start_date: str | None,
    end_date: str | None,
    event_type: str | None,
    search: str | None,
):
    """Shared query builder for list and export."""
    today_str = _today_pacific()

    q = (
        select(Event)
        .options(joinedload(Event.legislator))
        .where((Event.date >= today_str) | (Event.date.is_(None)))
    )

    if chamber and chamber != "all":
        q = q.join(Legislator).where(Legislator.chamber == chamber)
    else:
        q = q.join(Legislator)

    if start_date:
        q = q.where(Event.date >= start_date)
    if end_date:
        q = q.where(Event.date <= end_date)
    if event_type:
        q = q.where(Event.event_type == event_type)
    if search:
        pattern = f"%{search}%"
        q = q.where(
            Event.title.ilike(pattern)
            | Event.additional_details.ilike(pattern)
            | Event.address.ilike(pattern)
            | Legislator.name.ilike(pattern)
        )

    q = q.order_by(Event.date.asc(), Legislator.chamber, cast(Legislator.district, Integer))
    return q


@app.get("/api/events")
async def list_events(
    chamber: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    event_type: str | None = Query(None),
    search: str | None = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(100, ge=1, le=500),
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    q = _build_events_query(chamber, start_date, end_date, event_type, search)

    # Count total
    from sqlalchemy import func as sa_func

    count_q = select(sa_func.count()).select_from(q.subquery())
    total = (await session.execute(count_q)).scalar() or 0

    # Paginate
    offset = (page - 1) * per_page
    q = q.offset(offset).limit(per_page)
    result = await session.execute(q)
    events = result.scalars().unique().all()

    total_pages = max(1, -(-total // per_page))  # ceil division

    return {
        "events": [
            {
                "id": ev.id,
                "title": ev.title,
                "date": ev.date,
                "time": ev.time,
                "address": ev.address,
                "event_type": ev.event_type,
                "additional_details": ev.additional_details,
                "source_url": ev.source_url,
                "is_virtual": ev.is_virtual,
                "legislator_name": ev.legislator.name,
                "legislator_party": ev.legislator.party,
                "legislator_district": ev.legislator.district,
                "legislator_chamber": ev.legislator.chamber,
            }
            for ev in events
        ],
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages,
    }


def _format_date_human(date_str: str | None) -> str:
    """'2026-03-15' -> 'March 15, 2026'."""
    if not date_str:
        return ""
    try:
        d = date.fromisoformat(date_str)
        return d.strftime("%B %d, %Y").replace(" 0", " ")
    except ValueError:
        return date_str


def _format_time_human(time_str: str | None) -> str:
    """'18:00' -> '6:00 PM'."""
    if not time_str:
        return ""
    try:
        from datetime import datetime as dt

        t = dt.strptime(time_str, "%H:%M")
        return t.strftime("%I:%M %p").lstrip("0")
    except ValueError:
        return time_str


def _format_legislator_name(leg: Legislator) -> str:
    """Format as 'Assemblymember Jane Smith (D-42)' or 'Senator John Doe (R-15)'."""
    title = "Senator" if leg.chamber == "senate" else "Assemblymember"
    party_letter = leg.party[0] if leg.party else "?"
    return f"{title} {leg.name} ({party_letter}-{leg.district})"


@app.get("/api/events/export")
async def export_events(
    chamber: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    event_type: str | None = Query(None),
    search: str | None = Query(None),
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    q = _build_events_query(chamber, start_date, end_date, event_type, search)
    result = await session.execute(q)
    events = result.scalars().unique().all()

    from openpyxl.styles import Font

    wb = Workbook()
    headers = [
        "NAME",
        "DATE",
        "TIME",
        "ADDRESS",
        "TITLE OF EVENT",
        "EVENT LINK",
        "ADDITIONAL DETAILS",
    ]
    bold = Font(bold=True)

    assembly_events = [ev for ev in events if ev.legislator.chamber == "assembly"]
    senate_events = [ev for ev in events if ev.legislator.chamber == "senate"]

    def _fill_sheet(ws, ev_list):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold
        for ev in ev_list:
            name = _format_legislator_name(ev.legislator)
            details_parts = [ev.event_type or "", ev.additional_details or ""]
            details = " — ".join(p for p in details_parts if p)
            ws.append(
                [
                    name,
                    _format_date_human(ev.date),
                    _format_time_human(ev.time),
                    ev.address,
                    ev.title,
                    ev.source_url,
                    details,
                ]
            )
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    ws_assembly = wb.active
    ws_assembly.title = "Assembly"
    _fill_sheet(ws_assembly, assembly_events)

    ws_senate = wb.create_sheet("Senate")
    _fill_sheet(ws_senate, senate_events)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"ca_legislative_events_{_today_pacific()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
